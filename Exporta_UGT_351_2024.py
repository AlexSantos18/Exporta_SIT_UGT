import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle

# Carrega o arquivo Excel original
input_file = 'c:/sit/UGT_SIT_351_2024.xlsx'
df = pd.read_excel(input_file)

# Lista de colunas a excluir
colunas_excluir = ['COD_DESP', 'NR_ORÇAMENTO', 'DT_ORÇAMENTO', 'DT_PAGTO','DT_PAGTO_DEB', 'DS_PRODUTO']  

# Exclui as colunas especificadas
df_novo = df.drop(columns=colunas_excluir)

# Adiciona uma nova coluna com dados 
df_novo['coluna1']= ' '
df_novo['coluna2']= ' '  

# Lista de colunas na nova ordem desejada
nova_ordem_colunas = ['DATA', 'NF', 'RZ_SOCIAL', 'CNPJ', 'DESCRIÇÃO', 'coluna1', 'coluna2','NR_PAGTO', 'VALOR']  

# Reorganiza as colunas na nova ordem desejada
df_novo = df_novo[nova_ordem_colunas]

# Define o caminho para o novo arquivo Excel
output_file = 'c:/sit/exporta/exporta_ugt_351_2024.xlsx'

# Salva o novo DataFrame no novo arquivo Excel
df_novo.to_excel(output_file, index=False)

# Carrega o novo arquivo Excel para aplicar a formatação
wb = load_workbook(output_file)
ws = wb.active

# Cria um estilo de célula para data
date_style = NamedStyle(name="date_style", number_format='DD/MM/YYYY')

# Aplica o estilo de data à coluna 
for cell in ws['A']:
    cell.style = date_style
    cell.alignment = Alignment(horizontal="center") 

for cell in ws['B']:
    cell.alignment = Alignment(horizontal="center")

for cell in ws['H']:
    cell.alignment = Alignment(horizontal="center")

for cell in ws['D']:
    cell.alignment = Alignment(horizontal="center")
    cell.number_format = '00\.000\.000\/0000-00'

# Salva o arquivo Excel com a formatação aplicada
wb.save(output_file)

print("Novo arquivo Excel criado com a coluna formatada.")
print("Novo arquivo Excel criado, excluindo as colunas especificadas.")